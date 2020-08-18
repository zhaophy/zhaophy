#!/usr/bin/env python
# -*- coding: utf-8 -*-

from decimal import Decimal
import requests
import datetime, time
import json
import openpyxl
import datetime
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import os
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.styles import Font, colors, Alignment
import zipfile

HOST = '10.100.10.200'
USER = 'Admin'
PWD = 'nandou'
API = 'http://10.100.10.200/zabbix/api_jsonrpc.php'

today = datetime.date.today()

####创建当日日期的文件夹，用以将生成的EXCEL报表放入
def file():
    print os.getcwd()
    b = os.getcwd() + "\\" + str(today)
    try:
        os.mkdir(b)
    except:
        pass
file()


#获取ZABBIX登陆token
def get_zabbix_token():                                                         #定义获取zabbix登陆token方法
    data = {                                                                    #接口需要的参数
        "jsonrpc": "2.0",
        "method": "user.login",
        "params": {
            "user": USER,
            "password": PWD
        },
        "id": 1

    }
    url = 'http://10.100.10.200/zabbix/api_jsonrpc.php'                          #接口地址
    headers = {"Content-Type": "application/json"}                              #请求头数据
    response = requests.post(url, json.dumps(data), headers=headers).content    #封装数据以及请求头使用POST方法请求数据获取返回结果
    result = json.loads(response)['result']
    return result
token = get_zabbix_token()

###获取zabbix监控所有组
def get_group():
    data = {
    "jsonrpc": "2.0",
    "method": "hostgroup.get",
    "params": {
        "output": "extend",
    },
    "auth": token,
    "id": 1
    }
    url = API
    headers = {"Content-Type": "application/json"}
    gp_list = {}
    try:
        response = requests.post(url, json.dumps(data), headers=headers).content
        result = json.loads(response)['result']
       # print result
        for i in result:
            gp_list[i['name']] = i['groupid']

    except:
        gp_list = {}
    return gp_list

###把主机组作为输入参数，到处对应主机组的报表





##根据gid，获取该组的所有主机ID
def get_host(gid):
    data = {
    "jsonrpc": "2.0",
    "method": "host.get",
    "params": {
        "output": ["groupid", "name"],
        "groupids": gid
    },
    "auth": token,
    "id": 1
    }
    url = API
    headers = {"Content-Type": "application/json"}
    try:
        response = requests.post(url, json.dumps(data), headers=headers).content

        result = json.loads(response)['result']
    except:
        result = {}
    return result

##根据主机ID及keyID获取该监控项的值
def get_info(hostid,keyid):
    data = (
            {
                "jsonrpc": "2.0",
                "method": "item.get",
                "params": {
                    "output": "extend",
                    "hostids":hostid,  #这个是host的id号
                    "search":{
                        "key_":keyid,

                    }
                },
                "auth": token,
                "id": 1
            })
    url = API
    headers = {"Content-Type": "application/json"}
    try:
        response = requests.post(url, json.dumps(data), headers=headers).content
        k = json.loads(response)['result'][0]['lastvalue']
    except:
        k = 0
    return k
keys = ["system.cpu.load[percpu,avg5]","system.cpu.load[percpu,avg15]","system.cpu.util[,idle]"]



"""for i in gid_hid():
    #print i
    for j  in i:
        for m in keys:
            print get_info(j['hostid'],m)
"""


"""def get_group(hostid):
    data = {'jsonrpc': '2.0',
              'method': 'hostinterface.get',
              'params': {
                  'output': 'extend',
                  'hostids':hostid,
              },
              'auth': token,
              'id': 1
              }

    url = API
    headers = {"Content-Type": "application/json"}


    try:
        response = requests.post(url, json.dumps(data), headers=headers).content
        result = json.loads(response)['result'][0]['ip']
        return result
    except:
        return None
"""

##根据hostID获取其对应的IP地址
def get_ip(hostid):
    data = {'jsonrpc': '2.0',
              'method': 'hostinterface.get',
              'params': {
                  'output': 'extend',
                  'hostids':hostid,
              },
              'auth': token,
              'id': 1
              }

    url = API
    headers = {"Content-Type": "application/json"}


    try:
        response = requests.post(url, json.dumps(data), headers=headers).content
        result = json.loads(response)['result'][0]['ip']
        return result
    except:
        return 0


##定义字节、M的换算方法
def math(v):
    r = round((Decimal(v) / (1024 * 1024 * 1024)),2)
    return r

###获取监控项的值
def get_data(gid):
    d = []
    for i in get_host(gid):
        mem_total = math(get_info(i['hostid'],"vm.memory.size[total]"))
        mem_ava = math(get_info(i['hostid'],"vm.memory.size[available]"))
        mem_used = round((Decimal(mem_total) - Decimal(mem_ava)),2)
        if mem_total == 0:
            mem_used_percent = 0
        else:
            mem_used_percent = '{:.2%}'.format(Decimal(mem_ava) / Decimal(mem_total))

        disk_total = math(get_info(i['hostid'],"vfs.fs.size[/,total]"))
        disk_ava = math(get_info(i['hostid'],"vfs.fs.size[/,free]"))
        disk_used = round((Decimal(disk_total) - Decimal(disk_ava)),2)
        if disk_total == 0:
            disk_used_percent = 0
        else:
            disk_used_percent = '{:.2%}'.format(Decimal(disk_used) / Decimal(disk_total))
        data = [i['name'],get_ip(i['hostid']),mem_total,mem_ava,mem_used,mem_used_percent,disk_total,disk_ava,disk_used,disk_used_percent]
        for j in keys:
            data.append(get_info(i['hostid'],j))
        d.append(data)

    return d

#title = [[u'主机',u'IP',u'物理内存大小(G)',u'物理可用内存大小(G)',u'物理已使用内存大小(G)',u'根分区总大小(G)',u'根分区可用大小(G)',u'根分区已使用大小(G)',u'CPU5分钟负载',u'CPU15分钟负载',u'CPU平均空闲值']]
title = [[u'主机',u'IP',u'物理内存大小(G)',u'物理可用内存大小(G)',u'物理内存已使用大小(G)',u'物理内存已使用百分比',u'根分区总大小(G)',u'根分区可用大小(G)',u'根分区已使用大小(G)',u'根分区已使用百分比',u'CPU5分钟负载',u'CPU15分钟负载',u'CPU平均空闲值']]

#def to_excel(gid):
#    for i in get_data(gid):
  #      data.ap



def toexcel(input):
    #path = "C:\Users\ZHAOPHY\PycharmProjects\untitled\excel" + '/'
    path = os.getcwd() + '/' + str(today) + '/'
    gid = get_group()[input]
    data = title + get_data(gid)
    line = ['A','B','C','D','E','F','G','H','I','J','K','L','M']
    row = len(title)
    wb = openpyxl.Workbook()
    file_name = "%s_%s.xlsx" % (input,today)
    ws = wb.active
    ws.title = "REPORT"
    ws.sheet_properties.tabColor = '1072BA'
    fill = PatternFill("solid", fgColor="1ac6c0")
    for i in data:
        ws.append(i)

    for u in ws[1]:
        u.fill = fill
    for m in range(1,(len(data) + 1)):
        ws.row_dimensions[m].height = 18
    for n in line:
        ws.column_dimensions[n].width = 18

    wb.save(path + file_name)





gp = [u'ADAS 指南针',u'NewEnergy_Prodct',u'NewEnergy_Test',u'华神',u'华神国六尾气监管',u'商用车道路试验车',u'尾气排放',u'指南针武汉转发',u'楚道行',u'随专']

##压缩文件
def zip_file():
    for i in gp:
        toexcel(i)

    file_path = str(today)
    zipname = str(today)
    zip_name = file_path + '.zip'
    z = zipfile.ZipFile(zip_name,'w',zipfile.ZIP_DEFLATED)
    for dir_path, dir_names, file_names in os.walk(file_path):
        f_path = dir_path.replace(file_path, '')
        f_path = f_path and f_path + os.sep or ''
        for filename in file_names:
            z.write(os.path.join(dir_path, filename), f_path + filename)
    z.close()
    return zipname

zip_file()
##发送邮件
def SendEmail():

    path = str(today) + '.zip'
 #   print path
   # upath = unicode(str(path), 'utf-8')
  #  print path

    sender = 'nd6xwiki@163.com'
    senderPwd = 'nd6xrootroot'  # 邮箱密码我这里是放在环境变量中了（Win）
    receivers = ['3270387099@qq.com']  # 接收人邮箱
   # receivers = 'gongchaohui@dfmc.com.cn'
    # create a Instance
    TITLE = '监控报告_%s' % today
    message = MIMEMultipart()
    message['From'] = sender
    message['To'] = ",".join(receivers)
    subject = TITLE
    message['Subject'] = Header(subject, 'utf-8')
    # Message body content
    #MS = "您好:\n     附件为%s 资源使用情况报表，请查收" % (input）
    #message.attach(MIMEText(MS, 'plain', 'utf-8'))
    # Send xlsx file
    message.attach(MIMEText(' 您好, \n 请查收资源报表 \n\n Regards ', 'plain', 'utf-8'))
    att = MIMEText(open(path,'rb').read(), 'base64', 'utf-8')
    att["Content-Type"] = 'application/octet-stream'
    # Here you can rename the attachments in the message.
    att["Content-Disposition"] = 'attachment; filename="{}.zip"'.format(TITLE.encode('gbk'))
    message.attach(att)
    try:
        smtpObj = smtplib.SMTP('smtp.163.com','25')
        smtpObj.starttls()
        smtpObj.login(sender, senderPwd)  # 代理服务器帐号密码验证
        smtpObj.sendmail(sender, receivers, message.as_string())
        # terminating the session
        smtpObj.quit()
        print("Send email successful")
    except  smtplib.SMTPException as e:
        print(e.__doc__, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        print(e.__context__, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


SendEmail()